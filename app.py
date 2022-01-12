import pandas as pd
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side

def organize_pd_data(excel_name):
  df = pd.read_excel(excel_name, header=2)
  df.rename( columns={'日付/曜日':'日付'}, inplace=True )
  df.rename( columns={'Unnamed: 2':'曜日'}, inplace=True )
  df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
  df = df.drop(df.columns[2:5], axis=1)
  df = df.drop(df.columns[-4:],axis=1)
  df = df.drop(df.index[-3:])
  return df

def decimal2time(date, decimal):
  date = date.strftime('%Y/%m/%d')
  if type(decimal) == float:
    hour = str(int(decimal))
    minute = str(int((decimal % 1)*60))
  elif type(decimal) == str:
    hour = str(int(float(decimal)))
    minute = str(int((float(decimal) % 1)*60))
  else:
    print("typeをfloatかstrにしてください!")
  time = date + " " + hour + ":" + minute
  time = datetime.strptime(time, "%Y/%m/%d %H:%M")
  return time

def input_data_from_sheet(df):
  date_info_list = []
  for i in range(len(df)):
    data = df.loc[i]
    date_info = {}
    date_info['date'] = data['日付']
    date_info['day'] = data['曜日']
    opening = data['開館時間'].split('-')
    date_info['day_begin'] = decimal2time(data['日付'], opening[0])
    date_info['day_last'] = decimal2time(data['日付'], opening[1])
    date_info['members'] = []
    for j in range(len(df.columns)):
      if j < 3:
        continue
      else:
        if type(data[j]) == str:
          member = {}
          member['name'] = df.columns[j]
          worktime = data[j].split('-')
          member['begin'] = decimal2time(data['日付'], worktime[0])
          member['last'] = decimal2time(data['日付'], worktime[1])
          date_info['members'].append(member)
    date_info['members'] = sorted(date_info['members'], key=lambda x:x['begin'])
    date_info_list.append(date_info)
  return date_info_list

def data2excel(data):
  num = {0:'B', 1:'C', 2:'D', 3:'E', 4:'F', 5:'G', 6:'H'}
  side = Side(style='thin', color='000000')

  wb = openpyxl.Workbook()

  for i in range(len(data)):
    sheetname = data[i]['date'].strftime("%Y%m%d")
    wb.create_sheet(title=sheetname)
    ws = wb[sheetname]

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[3].height = 20

    ws['B1'] = data[i]['date'].strftime("%Y/%m/%d")
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C1'] = data[i]['day'] + "曜日"
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = "時刻"
    ws['A3'].border = Border(left=side, right=side, top=side, bottom=side)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    all_begin = 9.0
    all_begin = decimal2time(data[i]['date'], all_begin)
    members = data[i]['members']
    day_begin = data[i]['day_begin']
    day_last = data[i]['day_last']
    day_opening = (day_last - day_begin)/timedelta(minutes=30)

    for k in range(6):
      time = all_begin
      n=0

      if k >= len(members) and k < 6:
        ws.cell(row=3, column=2+k).border = Border(left=side, right=side, top=side, bottom=side)
        ws.column_dimensions["{}".format(num[k])].width = 13
        while time<day_last:
          time = time + timedelta(minutes=30)
          ws['{}{}'.format(num[k], 4+n)].border = Border(left=side, right=side, top=side, bottom=side)
          ws.row_dimensions[4+n].height = 26
          n+=1
      else:
        ws.cell(row=3, column=2+k).value = members[k]['name'] + "AA"
        ws.cell(row=3, column=2+k).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=3, column=2+k).border = Border(left=side, right=side, top=side, bottom=side)
        ws.column_dimensions["{}".format(num[k])].width = 13
        begin = members[k]['begin']
        last = members[k]['last']
        while time<begin:
          time = time + timedelta(minutes=30)
          ws['{}{}'.format(num[k], 4+n)].fill = PatternFill(patternType='darkGray', fgColor='909090')
          ws['{}{}'.format(num[k], 4+n)].border = Border(left=side, right=side, top=side, bottom=side)
          ws.row_dimensions[4+n].height = 26
          n+=1
        while time<last:
          time = time + timedelta(minutes=30)
          ws['{}{}'.format(num[k], 4+n)].border = Border(left=side, right=side, top=side, bottom=side)
          ws.row_dimensions[4+n].height = 26
          n+=1
        while time<day_last:
          time = time + timedelta(minutes=30)
          ws['{}{}'.format(num[k], 4+n)].fill = PatternFill(patternType='darkGray', fgColor='909090')
          ws['{}{}'.format(num[k], 4+n)].border = Border(left=side, right=side, top=side, bottom=side)
          ws.row_dimensions[4+n].height = 26
          n+=1

    time = all_begin
    n=0
    while time<day_last:
      ws.cell(row=4+n, column=1).value = time.strftime("%H:%M")
      ws.cell(row=4+n, column=1).border = Border(left=side, right=side, top=side, bottom=side)
      ws.cell(row=4+n, column=1).alignment = Alignment(horizontal='center', vertical='center')
      time = time + timedelta(minutes=30)
      n+=1

  wb.remove(wb.worksheets[0])

  Name = "体制表" + data[0]['date'].strftime("%Y%m")
  file_xlsx = Name + ".xlsx"
  
  wb.save(file_xlsx)

def main():
    filename = input("シフト表のエクセルファイルのパスを入力してください。")
    df = organize_pd_data(filename)
    date_info_list = input_data_from_sheet(df)
    data2excel(date_info_list)

if __name__ == "__main__":
    main()